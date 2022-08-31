import os
import openpyxl
from openpyxl import load_workbook

path = input("Input file name") + ".xlsx"
a = os.listdir(rf"{path}")
q = []

try:
    for i in range(len(a)):
        s = a[i].strip()
        wb = load_workbook(filename=rf"{path}\your_file.xlsx")
        w = wb.active
        q.append([w.cell(1, i ).value for i in range (10)])
except Exception as ex:
    pass
f = openpyxl.Workbook()
fs = f.active
ff = f.create_sheet()
for i in range(1, len(q)+1):
    for r in range(1,7):
        fff = ff.cell(i, r, q[i-1][r-1])

f.save(rf"{path}\your_file.xlsx")
